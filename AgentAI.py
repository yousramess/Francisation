"""
Agent IA — Convertir PDF en Excel
Requires: pip install anthropic pdfplumber openpyxl
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import anthropic
from openai import OpenAI

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
API_KEY = "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"   # <-- Remplacez par votre clé Anthropic
# ──────────────────────────────────────────────────────────────────────────────


def extraire_texte_pdf(chemin_pdf: str) -> str:
    """Extrait tout le texte d'un PDF (texte + tableaux)."""
    texte = ""
    with pdfplumber.open(chemin_pdf) as pdf:
        for i, page in enumerate(pdf.pages):
            # Essai extraction de tableau structuré
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row:
                            texte += " | ".join(str(c) if c else "" for c in row) + "\n"
            else:
                # Sinon texte brut
                t = page.extract_text()
                if t:
                    texte += t + "\n"
    return texte.strip()


#def appeler_claude(texte_pdf: str, client: anthropic.Anthropic) -> list[dict]:
    """Envoie le texte à Claude et retourne les données structurées."""
    system_prompt = """Tu es un assistant spécialisé dans l'extraction de données depuis des tableaux PDF.

À partir du texte fourni, extrais toutes les personnes trouvées.
Pour chaque personne:
- Extrais: nom_complet, courriel, telephone (depuis le PDF)
- Sépare: nom et prenom à partir du nom_complet (nom = dernier mot en majuscule ou dernier mot, prenom = reste)
- Laisse à vide (chaîne vide ""): francisation, groupe, odoo

Réponds UNIQUEMENT avec un tableau JSON valide, sans texte ni balises markdown autour.
Format exact:
[
  {
    "nom_complet": "...",
    "nom": "...",
    "prenom": "...",
    "courriel": "...",
    "telephone": "...",
    "francisation": "",
    "groupe": "",
    "odoo": ""
  }
]

Si aucune donnée trouvée, retourne exactement: []"""

def appeler_chatgpt(texte, client):
    prompt = f"""
    Extrais les informations importantes du texte suivant et retourne un JSON structuré.
    Chaque personne doit contenir: nom, prenom, email, telephone.

    Texte:
    {texte}
    """

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "Tu es un assistant qui extrait des données structurées."},
            {"role": "user", "content": prompt}
        ],
        temperature=0
    )

    import json
    contenu = response.choices[0].message.content

    try:
        return json.loads(contenu)
    except:
        print("Erreur JSON :", contenu)
        return None

    reponse = message.content[0].text.strip()
    # Nettoyer les éventuelles balises markdown
    reponse = reponse.replace("```json", "").replace("```", "").strip()
    return json.loads(reponse)


def creer_excel(donnees: list[dict], chemin_sortie: str):
    """Crée un fichier Excel formaté avec les données."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"

    # Couleurs
    BLEU_HEADER = "1F4E79"
    BLEU_CLAIR  = "D6E4F0"
    GRIS_CLAIR  = "F2F2F2"

    en_tetes = ["Nom, Prénom", "Nom", "Prénom", "Courriel",
                "Téléphone", "Francisation", "Groupe", "Odoo"]
    cles     = ["nom, Prénom", "nom", "prenom", "courriel",
                "telephone", "francisation", "groupe", "odoo"]
    largeurs = [28, 18, 18, 30, 18, 18, 15, 15]

    # Style en-têtes
    font_header = Font(bold=True, color="FFFFFF", size=11)
    fill_header = PatternFill("solid", fgColor=BLEU_HEADER)
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    ws.row_dimensions[1].height = 28
    for col_idx, (titre, largeur) in enumerate(zip(en_tetes, largeurs), start=1):
        cell = ws.cell(row=1, column=col_idx, value=titre)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = largeur

    # Données
    for row_idx, personne in enumerate(donnees, start=2):
        fill_row = PatternFill("solid", fgColor=GRIS_CLAIR if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, cle in enumerate(cles, start=1):
            val = personne.get(cle, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill_row
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 20

    # Figer la ligne d'en-tête
    ws.freeze_panes = "A2"

    # Filtre automatique
    ws.auto_filter.ref = ws.dimensions

    wb.save(chemin_sortie)


# ─── INTERFACE GRAPHIQUE ───────────────────────────────────────────────────────

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Agent IA — PDF vers Excel")
        self.geometry("620x540")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")

        self.chemin_pdf = tk.StringVar()
        self.chemin_excel = tk.StringVar()
        self._construire_ui()

    def _construire_ui(self):
        # Titre
        tk.Label(self, text="Agent IA — PDF → Excel",
                 font=("Segoe UI", 18, "bold"), bg="#1F4E79", fg="white",
                 pady=16).pack(fill="x")

        tk.Label(self, text="Convertit votre tableau PDF en fichier Excel enrichi",
                 font=("Segoe UI", 10), bg="#F0F4F8", fg="#555").pack(pady=(10, 4))

        cadre = tk.Frame(self, bg="#F0F4F8", padx=30)
        cadre.pack(fill="x", pady=10)

        # --- Clé API ---
        tk.Label(cadre, text="Clé API Anthropic :", font=("Segoe UI", 9, "bold"),
                 bg="#F0F4F8", anchor="w").grid(row=0, column=0, sticky="w", pady=(0,2))
        self.champ_cle = tk.Entry(cadre, width=55, show="*", font=("Segoe UI", 9),
                                  relief="solid", bd=1)
        self.champ_cle.insert(0, API_KEY if API_KEY != "sk-ant-VOTRE_CLE_ICI" else "")
        self.champ_cle.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0,12))
        tk.Label(cadre, text="(Obtenez votre clé sur console.anthropic.com)",
                 font=("Segoe UI", 8), bg="#F0F4F8", fg="#888", anchor="w"
                 ).grid(row=2, column=0, sticky="w", pady=(0,14))

        # --- Fichier PDF ---
        tk.Label(cadre, text="Fichier PDF à convertir :", font=("Segoe UI", 9, "bold"),
                 bg="#F0F4F8", anchor="w").grid(row=3, column=0, sticky="w", pady=(0,2))
        champ_pdf = tk.Entry(cadre, textvariable=self.chemin_pdf, width=44,
                             font=("Segoe UI", 9), relief="solid", bd=1, state="readonly")
        champ_pdf.grid(row=4, column=0, sticky="ew", pady=(0,10))
        tk.Button(cadre, text="Parcourir…", command=self._choisir_pdf,
                  font=("Segoe UI", 9), bg="#2980B9", fg="white",
                  relief="flat", padx=10, cursor="hand2"
                  ).grid(row=4, column=1, padx=(8,0), pady=(0,10))

        # --- Fichier Excel sortie ---
        tk.Label(cadre, text="Enregistrer le Excel sous :", font=("Segoe UI", 9, "bold"),
                 bg="#F0F4F8", anchor="w").grid(row=5, column=0, sticky="w", pady=(0,2))
        champ_xl = tk.Entry(cadre, textvariable=self.chemin_excel, width=44,
                            font=("Segoe UI", 9), relief="solid", bd=1, state="readonly")
        champ_xl.grid(row=6, column=0, sticky="ew", pady=(0,20))
        tk.Button(cadre, text="Choisir…", command=self._choisir_excel,
                  font=("Segoe UI", 9), bg="#2980B9", fg="white",
                  relief="flat", padx=10, cursor="hand2"
                  ).grid(row=6, column=1, padx=(8,0), pady=(0,20))

        cadre.columnconfigure(0, weight=1)

        # Bouton principal
        self.btn_convertir = tk.Button(
            self, text="▶  Convertir en Excel",
            font=("Segoe UI", 12, "bold"), bg="#1F4E79", fg="white",
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._lancer_conversion
        )
        self.btn_convertir.pack(pady=4)

        # Barre de progression
        self.progress = ttk.Progressbar(self, length=500, mode="indeterminate")
        self.progress.pack(pady=10)

        # Journal
        self.journal = tk.Text(self, height=7, width=72, font=("Consolas", 9),
                               bg="#1a1a2e", fg="#00e5ff", relief="flat",
                               state="disabled", wrap="word", padx=8, pady=6)
        self.journal.pack(padx=30, pady=(0,14))

    def _choisir_pdf(self):
        chemin = filedialog.askopenfilename(
            title="Choisir un fichier PDF",
            filetypes=[("Fichiers PDF", "*.pdf")]
        )
        if chemin:
            self.chemin_pdf.set(chemin)
            # Proposer automatiquement le même dossier pour la sortie
            dossier = os.path.dirname(chemin)
            nom = os.path.splitext(os.path.basename(chemin))[0]
            self.chemin_excel.set(os.path.join(dossier, f"{nom}_converti.xlsx"))

    def _choisir_excel(self):
        chemin = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")]
        )
        if chemin:
            self.chemin_excel.set(chemin)

    def _log(self, msg: str):
        self.journal.config(state="normal")
        self.journal.insert("end", msg + "\n")
        self.journal.see("end")
        self.journal.config(state="disabled")

    def _lancer_conversion(self):
        cle = self.champ_cle.get().strip()
        pdf = self.chemin_pdf.get().strip()
        xl  = self.chemin_excel.get().strip()

        if not cle:
            messagebox.showerror("Clé manquante", "Veuillez entrer votre clé API Anthropic.")
            return
        if not pdf:
            messagebox.showerror("PDF manquant", "Veuillez choisir un fichier PDF.")
            return
        if not xl:
            messagebox.showerror("Destination manquante", "Veuillez choisir où enregistrer le fichier Excel.")
            return

        self.btn_convertir.config(state="disabled")
        self.progress.start(12)
        self.journal.config(state="normal")
        self.journal.delete("1.0", "end")
        self.journal.config(state="disabled")

        threading.Thread(target=self._convertir, args=(cle, pdf, xl), daemon=True).start()

    def _convertir(self, cle: str, pdf: str, xl: str):
     try:
        self._log("📄 Lecture du fichier PDF…")
        texte = extraire_texte_pdf(pdf)
        if not texte:
            raise ValueError("Le PDF semble vide ou ne contient pas de texte lisible.")
        self._log(f"✅ {len(texte)} caractères extraits.")

        self._log("🤖 Envoi à ChatGPT pour analyse…")
        client = OpenAI(api_key=cle)
        donnees = appeler_chatgpt(texte, client)

        if not donnees:
            raise ValueError("Aucune donnée trouvée dans le PDF. Vérifiez le contenu.")

        self._log(f"✅ {len(donnees)} personne(s) trouvée(s).")

        self._log("📊 Création du fichier Excel…")
        creer_excel(donnees, xl)
        self._log(f"✅ Fichier enregistré : {xl}")
        self._log("─" * 50)
        self._log("🎉 Conversion terminée avec succès !")

        self.after(0, lambda: messagebox.showinfo(
            "Succès !",
            f"Conversion réussie !\n\n{len(donnees)} personne(s) exportée(s).\n\nFichier : {xl}"
         ))

     except json.JSONDecodeError:
            self._log("❌ Erreur : L'IA n'a pas retourné un JSON valide.")
            self.after(0, lambda: messagebox.showerror("Erreur", "L'IA n'a pas pu analyser le tableau. Essayez un autre PDF."))
     #except Exception as e:
       #     self._log(f"❌ Erreur : {e}")
        #    self.after(0, lambda: messagebox.showerror("Erreur", str(e)))
     except Exception as e:
       err = str(e)
       self.after(0, self._afficher_erreur, err)
     finally:
            self.after(0, self._fin_conversion)
    
    def _afficher_erreur(self, err):
        messagebox.showerror("Erreur", err)
        self._log(f"❌ Erreur : {err}")

    def _fin_conversion(self):
        self.progress.stop()
        self.btn_convertir.config(state="normal")


if __name__ == "__main__":
    app = Application()
    app.mainloop()